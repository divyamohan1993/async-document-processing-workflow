import csv
import io
import json
import os
import re
from collections import Counter


STOPWORDS = {
    "the", "a", "an", "and", "or", "but", "in", "on", "at", "to", "for",
    "of", "with", "by", "from", "is", "was", "are", "were", "be", "been",
    "being", "have", "has", "had", "do", "does", "did", "will", "would",
    "could", "should", "may", "might", "shall", "can", "need", "dare",
    "it", "its", "this", "that", "these", "those", "i", "you", "he", "she",
    "we", "they", "me", "him", "her", "us", "them", "my", "your", "his",
    "our", "their", "mine", "yours", "hers", "ours", "theirs", "what",
    "which", "who", "whom", "where", "when", "why", "how", "all", "each",
    "every", "both", "few", "more", "most", "other", "some", "such", "no",
    "nor", "not", "only", "own", "same", "so", "than", "too", "very",
    "just", "because", "as", "until", "while", "if", "then", "about",
    "up", "out", "into", "through", "during", "before", "after", "above",
    "below", "between", "under", "again", "further", "once", "here",
    "there", "also", "new", "like", "well", "even", "back", "any",
    "much", "now", "make", "made", "get", "got", "go", "went", "come",
    "came", "take", "took", "see", "saw", "know", "knew", "think",
    "thought", "say", "said", "tell", "told", "one", "two", "first",
}


def extract_text(file_path: str, file_type: str) -> str:
    """Extract text content from a file based on its type."""
    file_type = file_type.lower().lstrip(".")

    if file_type == "txt" or file_type == "md" or file_type == "rtf":
        return _extract_text_file(file_path)
    elif file_type == "pdf":
        return _extract_pdf(file_path)
    elif file_type == "docx":
        return _extract_docx(file_path)
    elif file_type == "csv":
        return _extract_csv(file_path)
    elif file_type == "xlsx":
        return _extract_xlsx(file_path)
    elif file_type == "json":
        return _extract_json(file_path)
    elif file_type == "xml":
        return _extract_xml(file_path)
    else:
        return _extract_binary(file_path)


def _extract_text_file(file_path: str) -> str:
    encodings = ["utf-8", "latin-1", "cp1252", "ascii"]
    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    return _extract_binary(file_path)


def _extract_pdf(file_path: str) -> str:
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(file_path)
        text_parts = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n".join(text_parts)
    except Exception as e:
        raise RuntimeError(f"Failed to extract PDF text: {e}")


def _extract_docx(file_path: str) -> str:
    try:
        from docx import Document

        doc = Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        raise RuntimeError(f"Failed to extract DOCX text: {e}")


def _extract_csv(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = []
            for row in reader:
                rows.append(" | ".join(row))
            return "\n".join(rows)
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            reader = csv.reader(f)
            rows = []
            for row in reader:
                rows.append(" | ".join(row))
            return "\n".join(rows)


def _extract_xlsx(file_path: str) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                cell_values = [str(cell) if cell is not None else "" for cell in row]
                text_parts.append(" | ".join(cell_values))
        wb.close()
        return "\n".join(text_parts)
    except Exception as e:
        raise RuntimeError(f"Failed to extract XLSX text: {e}")


def _extract_json(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return json.dumps(data, indent=2, default=str)


def _extract_xml(file_path: str) -> str:
    """Extract text content from XML, stripping tags."""
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()
    # Strip XML tags to get text content
    text = re.sub(r"<[^>]+>", " ", content)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _extract_binary(file_path: str) -> str:
    with open(file_path, "rb") as f:
        raw = f.read()
    try:
        return raw.decode("utf-8", errors="replace")
    except Exception:
        return raw.decode("latin-1", errors="replace")


def generate_summary(text: str, max_length: int = 200) -> str:
    """Generate a summary from text content."""
    if not text or not text.strip():
        return "No text content available."

    cleaned = re.sub(r"\s+", " ", text).strip()

    if len(cleaned) <= max_length:
        return cleaned

    # Try to truncate at sentence boundary
    truncated = cleaned[:max_length + 50]
    sentence_ends = [m.end() for m in re.finditer(r"[.!?]\s", truncated)]
    if sentence_ends:
        # Find the last sentence end before max_length
        valid_ends = [e for e in sentence_ends if e <= max_length + 20]
        if valid_ends:
            return truncated[: valid_ends[-1]].strip()

    # Truncate at word boundary
    truncated = cleaned[:max_length]
    last_space = truncated.rfind(" ")
    if last_space > max_length // 2:
        return truncated[:last_space].strip() + "..."
    return truncated.strip() + "..."


def extract_keywords(text: str, max_keywords: int = 10) -> list[str]:
    """Extract significant keywords from text."""
    if not text or not text.strip():
        return []

    # Tokenize: extract words
    words = re.findall(r"\b[a-zA-Z]{3,}\b", text.lower())

    # Filter stopwords
    meaningful = [w for w in words if w not in STOPWORDS and len(w) >= 3]

    # Count frequencies
    counter = Counter(meaningful)

    # Return top keywords
    return [word for word, _ in counter.most_common(max_keywords)]


def determine_category(file_type: str, text: str) -> str:
    """Categorize document based on file type and content analysis."""
    file_type = file_type.lower().lstrip(".")
    text_lower = text.lower() if text else ""

    if file_type in ("csv", "xlsx"):
        return "Spreadsheet Data"

    # Check for invoice patterns
    invoice_patterns = [
        r"\binvoice\b", r"\bbill\s+to\b", r"\btotal\s+due\b",
        r"\bpayment\b.*\bdue\b", r"\binv[\s\-#]", r"\bamount\s+due\b",
    ]
    for pattern in invoice_patterns:
        if re.search(pattern, text_lower):
            return "Invoice"

    # Check for letter patterns
    letter_patterns = [
        r"\bdear\s+\w+", r"\bsincerely\b", r"\bregards\b",
        r"\byours\s+(truly|faithfully)\b", r"\bto\s+whom\s+it\s+may\s+concern\b",
    ]
    for pattern in letter_patterns:
        if re.search(pattern, text_lower):
            return "Letter"

    # Check for report patterns
    report_patterns = [
        r"\bexecutive\s+summary\b", r"\bconclusion\b", r"\bfindings\b",
        r"\banalysis\b", r"\breport\b", r"\bintroduction\b.*\bmethodology\b",
    ]
    for pattern in report_patterns:
        if re.search(pattern, text_lower):
            return "Report"

    # Check for technical document patterns
    tech_patterns = [
        r"\bapi\b", r"\bfunction\b", r"\bclass\b", r"\bmodule\b",
        r"\bimport\b", r"\bdef\b", r"\breturn\b", r"\bspecification\b",
        r"\barchitecture\b", r"\bimplementation\b",
    ]
    tech_count = sum(1 for p in tech_patterns if re.search(p, text_lower))
    if tech_count >= 3:
        return "Technical Document"

    return "General Document"


def extract_title(text: str, filename: str) -> str:
    """Extract a title from text content or derive from filename."""
    if text and text.strip():
        first_line = text.strip().split("\n")[0].strip()
        # Clean up the first line
        first_line = re.sub(r"^[#=\-*]+\s*", "", first_line).strip()
        if first_line and len(first_line) <= 200 and len(first_line) >= 3:
            return first_line

    # Derive from filename
    name = os.path.splitext(os.path.basename(filename))[0]
    name = re.sub(r"[_\-]+", " ", name)
    name = re.sub(r"[^\w\s]", "", name)
    return name.strip().title() or "Untitled Document"
